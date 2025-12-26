import openpyxl
import os

print("="*70)
print("📁 EXCEL 폴더 vs GENERATED_EXCELS 폴더 비교")
print("="*70)

# 원본 템플릿
templates = {
    'I': 'excel/1. 지중경사계(한남동).xlsx',
    'W': 'excel/2. 지하수위계(한남동).xlsx',
    'S': 'excel/3. 변형률계(1~7 3단샘플).xlsx',
    'C': 'excel/4. 균열측정계(한남동).xlsx',
    'T': 'excel/5. 건물경사계 (한남동).xlsx',
    'SE': 'excel/6. 지표침하계(한남동).xlsx'
}

# 생성된 파일들 찾기
generated_folder = 'generated_excels'
generated_files = {}
if os.path.exists(generated_folder):
    for filename in os.listdir(generated_folder):
        if filename.endswith('.xlsx'):
            filepath = os.path.join(generated_folder, filename)
            generated_files[filename] = filepath

print(f"\n🔍 원본 템플릿: {len(templates)}개")
print(f"🔍 생성된 파일: {len(generated_files)}개")
print("\n" + "="*70)

# 각 센서별 비교
sensor_names = {
    'I': '지중경사계',
    'W': '지하수위계',
    'S': '변형률계',
    'C': '균열측정계',
    'T': '건물경사계',
    'SE': '지표침하계'
}

for code, name in sensor_names.items():
    template_path = templates[code]
    
    print(f"\n{'─'*70}")
    print(f"📊 [{code}] {name}")
    print(f"{'─'*70}")
    
    # 원본 분석
    if os.path.exists(template_path):
        wb_orig = openpyxl.load_workbook(template_path)
        file_size = os.path.getsize(template_path)
        
        print(f"\n✅ 원본: {os.path.basename(template_path)}")
        print(f"   크기: {file_size:,} bytes ({file_size/1024:.1f} KB)")
        print(f"   시트: {len(wb_orig.sheetnames)}개 - {wb_orig.sheetnames[:5]}" + 
              (f"..." if len(wb_orig.sheetnames) > 5 else ""))
        
        wb_orig.close()
    else:
        print(f"❌ 원본 없음")
    
    # 생성된 파일 찾기
    matching_generated = [f for f in generated_files.keys() if name in f]
    
    if matching_generated:
        for gen_filename in matching_generated:
            gen_path = generated_files[gen_filename]
            wb_gen = openpyxl.load_workbook(gen_path)
            gen_size = os.path.getsize(gen_path)
            
            print(f"\n🔧 생성: {gen_filename}")
            print(f"   크기: {gen_size:,} bytes ({gen_size/1024:.1f} KB)")
            print(f"   시트: {len(wb_gen.sheetnames)}개 - {wb_gen.sheetnames[:5]}" + 
                  (f"..." if len(wb_gen.sheetnames) > 5 else ""))
            
            # 크기 차이
            if os.path.exists(template_path):
                size_diff = gen_size - file_size
                print(f"   📏 크기 차이: {size_diff:+,} bytes ({size_diff/1024:+.1f} KB)")
            
            wb_gen.close()
    else:
        print(f"\n⚠️  생성된 파일 없음")

print("\n" + "="*70)
print("✅ 비교 완료")
