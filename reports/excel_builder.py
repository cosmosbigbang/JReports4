import pandas as pd
import os
import math
import random
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def apply_style(ws, min_row, max_row, min_col, max_col):
    """지정 범위에 테두리와 가운데 정렬 적용"""
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = align

def set_col_width(ws, col_map):
    """열 너비 설정 (예: {'A': 15, 'B': 10})"""
    for col_letter, width in col_map.items():
        ws.column_dimensions[col_letter].width = width

def create_excel_files(sensor_counts, project_info, base_dir="generated_excels"):
    output_dir = os.path.join(base_dir, project_info)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    created_log = []

    # ==============================================================================
    # 1. 지중경사계 (I) - [완벽 구현]
    # ==============================================================================
    if sensor_counts.get('I', 0) > 0:
        count = sensor_counts['I']
        filename = f"지중경사계({project_info}).xlsx"
        filepath = os.path.join(output_dir, filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for i in range(1, count + 1):
            ws = wb.create_sheet(title=f"{i}data")
            
            # [Title & Metadata]
            ws['A1'] = "지중경사계 DATA"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:F1') # 제목 병합
            
            metadata = [
                (4, f"■  현장명(Site) : {project_info} 근린생활시설 신축공사"),
                (5, f"■  설치위치(Location) : I-{i}"),
                (6, "■  계기종류(Type) : INCLINOMETER"),
                (7, "■  심도(Depth) : GL.(-) 10.0m")
            ]
            for r, text in metadata:
                ws[f'A{r}'] = text
                ws.merge_cells(f'A{r}:F{r}') # 메타데이터 넓게 병합
                ws[f'A{r}'].alignment = Alignment(horizontal='left')

            ws['O7'] = "(단위 : mm)"

            # [Header] 9-10행
            ws['A9'] = "Depth\n(GL.-m)"
            ws.merge_cells('A9:A10')
            ws['A9'].font = Font(bold=True)
            
            # 날짜 & 방향 (2회차 예시)
            dates = ["2025-11-10", "2025-11-13"]
            start_col = 3 # C열
            for dt in dates:
                ws.merge_cells(start_row=9, start_column=start_col, end_row=9, end_column=start_col+1)
                ws.cell(row=9, column=start_col, value=dt).font = Font(bold=True)
                ws.cell(row=10, column=start_col, value="AB방향").font = Font(bold=True)
                ws.cell(row=10, column=start_col+1, value="CD방향").font = Font(bold=True)
                start_col += 2
            
            apply_style(ws, 9, 10, 1, start_col-1)
            set_col_width(ws, {'A': 12, 'B': 2})

            # [Sample Data]
            for r in range(5):
                row_num = 11 + r
                ws.cell(row=row_num, column=1, value=0.5 * (r + 1))
                curr_col = 3
                for _ in dates:
                    ws.cell(row=row_num, column=curr_col, value=round(random.uniform(-0.5, 0.5), 2))
                    ws.cell(row=row_num, column=curr_col+1, value=round(random.uniform(-0.5, 0.5), 2))
                    curr_col += 2
            apply_style(ws, 11, 15, 1, start_col-1)

        wb.save(filepath)
        created_log.append(f"✅ {filename}")

    # ==============================================================================
    # 2. 지하수위계 (W) - [완벽 구현]
    # ==============================================================================
    if sensor_counts.get('W', 0) > 0:
        count = sensor_counts['W']
        filename = f"지하수위계({project_info}).xlsx"
        filepath = os.path.join(output_dir, filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for i in range(1, count + 1):
            ws = wb.create_sheet(title=f"W-{i}")
            
            ws['A1'] = "지 하 수 위 계 DATA"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:D1')

            meta = [
                (4, f"■ 현장명 : {project_info}"),
                (5, f"■ 설치위치 : W-{i}"),
                (6, "■ 계기종류 : Water Levelmeter"),
                (7, "■ 공식 : 수위차(m) = 측정치 - 초기치")
            ]
            for r, t in meta:
                ws[f'A{r}'] = t
                ws.merge_cells(f'A{r}:D{r}')
                ws[f'A{r}'].alignment = Alignment(horizontal='left')

            headers = ["Date", "측정치(m)", "수위차(m)", "비 고"]
            for c, text in enumerate(headers, 1):
                cell = ws.cell(row=14, column=c, value=text)
                cell.font = Font(bold=True)
            apply_style(ws, 14, 14, 1, 4)
            set_col_width(ws, {'A': 15, 'B': 12, 'C': 12, 'D': 20})

            # Data
            init_val = -10.0
            for r in range(5):
                row_num = 15 + r
                cur = init_val + random.uniform(-0.5, 0.5)
                ws.cell(row=row_num, column=1, value=f"2025-01-{r+1:02d}")
                ws.cell(row=row_num, column=2, value=round(cur, 2))
                ws.cell(row=row_num, column=3, value=round(cur - init_val, 2))
                if r==0: ws.cell(row=row_num, column=4, value="초기치")
            apply_style(ws, 15, 19, 1, 4)

        wb.save(filepath)
        created_log.append(f"✅ {filename}")

    # ==============================================================================
    # 3. 변형률계 (S) - [3단 헤더 완벽 구현]
    # ==============================================================================
    if sensor_counts.get('S', 0) > 0:
        count = sensor_counts['S']
        filename = f"변형률계({project_info}).xlsx"
        filepath = os.path.join(output_dir, filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        locs = math.ceil(count / 3)
        for i in range(1, locs + 1):
            ws = wb.create_sheet(title=f"S{i}")
            ws['A1'] = "Strain Gauge Data Sheet"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:N1')
            
            ws['A3'] = f"현 장 명 : {project_info}"
            ws.merge_cells('A3:N3')
            ws['A4'] = f"설 치 위 치 : S{i}-1,2,3 - STRUT"
            ws.merge_cells('A4:N4')

            # Row 12 (Main Header)
            ws['A12'] = "측정일"
            ws.merge_cells('A12:A14')
            
            headers = [("B12","D12","측정치(µε)"), ("E12","G12","응력(kg/㎠)"), 
                       ("H12","J12","응력(Mpa)"), ("K12","M12","축력(Ton)")]
            for s, e, t in headers:
                ws[s] = t
                ws.merge_cells(f"{s}:{e}")
                ws[s].font = Font(bold=True)
            
            ws['N12'] = "비 고"
            ws.merge_cells('N12:N14')

            # Row 13 (Layers)
            sub = ["1단", "2단", "3단"] * 4
            for idx, val in enumerate(sub):
                ws.cell(row=13, column=2+idx, value=val).font = Font(bold=True)

            # Row 14 (Sensor IDs)
            ids = [f"S{i}-1", f"S{i}-2", f"S{i}-3"] * 4
            for idx, val in enumerate(ids):
                ws.cell(row=14, column=2+idx, value=val).font = Font(bold=True)

            apply_style(ws, 12, 14, 1, 14)
            set_col_width(ws, {'A': 15, 'N': 15})

            # Data
            for r in range(5):
                row = 15 + r
                ws.cell(row=row, column=1, value=f"2025-01-{r+1:02d}")
                for c in range(2, 14):
                    ws.cell(row=row, column=c, value=int(random.uniform(2800, 3100)))
            apply_style(ws, 15, 19, 1, 14)

        wb.save(filepath)
        created_log.append(f"✅ {filename}")

    # ==============================================================================
    # 4. 균열측정계 (C) - [좌우 병렬 표 완벽 구현]
    # ==============================================================================
    if sensor_counts.get('C', 0) > 0:
        count = sensor_counts['C']
        filename = f"균열측정계({project_info}).xlsx"
        filepath = os.path.join(output_dir, filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for i in range(1, count + 1):
            ws = wb.create_sheet(title=f"C-{i}")
            ws['A1'] = "균 열 측 정 계 DATA"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:I1')
            
            # 우측 상단 라벨
            ws['K1'] = f"CK-{i}(상하변위)" 
            ws['K4'] = f"CK-{i}(좌우변위)"

            # 메타데이터
            meta = [
                (4, "■ 계측관리업체 : ㈜지오테크이앤씨"),
                (5, f"■ 현장명 : {project_info}"),
                (6, f"■ 설치위치 : CK-{i}"),
                (7, "■ 계기종류 : Multi Crack Gauge"),
                (8, "■ 공식 : 변위량 = 금회 - 초기")
            ]
            for r, t in meta:
                ws[f'A{r}'] = t
                ws.merge_cells(f'A{r}:F{r}')
                ws[f'A{r}'].alignment = Alignment(horizontal='left')

            # Headers (Row 10)
            l_h = ["Date", "측정치(상하)", "변위량", "비 고"]
            r_h = ["Date", "측정치(좌우)", "변위량", "비 고"]
            
            for idx, t in enumerate(l_h, 1):
                ws.cell(row=10, column=idx, value=t).font = Font(bold=True)
            for idx, t in enumerate(r_h, 6): # F열=6
                ws.cell(row=10, column=idx, value=t).font = Font(bold=True)
            
            apply_style(ws, 10, 10, 1, 4) # 좌측 헤더
            apply_style(ws, 10, 10, 6, 9) # 우측 헤더
            set_col_width(ws, {'A':12, 'B':12, 'C':12, 'D':10, 'E':2, 'F':12, 'G':12, 'H':12, 'I':10})

            # Data
            for r in range(5):
                rn = 11 + r
                ws.cell(row=rn, column=1, value=f"2025-01-{r+1:02d}")
                ws.cell(row=rn, column=6, value=f"2025-01-{r+1:02d}")
                for c in [2,3,7,8]:
                    ws.cell(row=rn, column=c, value=0.0)
            apply_style(ws, 11, 15, 1, 4)
            apply_style(ws, 11, 15, 6, 9)

        wb.save(filepath)
        created_log.append(f"✅ {filename}")

    # ==============================================================================
    # 5. 건물경사계 (T) - [4단 헤더 및 단위 행 완벽 구현]
    # ==============================================================================
    if sensor_counts.get('T', 0) > 0:
        count = sensor_counts['T']
        filename = f"건물경사계({project_info}).xlsx"
        filepath = os.path.join(output_dir, filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for i in range(1, count + 1):
            ws = wb.create_sheet(title=f"T-{i}")
            ws['A1'] = "건물경사계 DATA"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:K1')
            
            meta = [
                (4, "■ 계측관리업체 : ㈜지오테크이앤씨"),
                (5, f"■ 현장명 : {project_info}"),
                (6, f"■ 계측기번호 : T-{i}"),
                (7, "■ 계기종류 : TILTMETER"),
                (8, "■ 공식 : 변위량 = L × Sinθ")
            ]
            for r, t in meta:
                ws[f'A{r}'] = t
                ws.merge_cells(f'A{r}:F{r}')
                ws[f'A{r}'].alignment = Alignment(horizontal='left')

            # Headers (13-16행)
            ws['B13']="Data Reading"; ws.merge_cells('B13:C13')
            ws['D13']="Displace"; ws.merge_cells('D13:F13')
            ws['G13']="Data Reading"; ws.merge_cells('G13:H13')
            ws['I13']="Displace"; ws.merge_cells('I13:K13')
            
            ws.cell(row=14, column=2, value='AB')
            ws.cell(row=14, column=3, value='BA')
            ws.cell(row=14, column=7, value='CD')
            ws.cell(row=14, column=8, value='DC')

            h_15 = ["계측일자", "AB측정치", "BA측정치", "(AB-BA)/2", "변위량", "길이환산", 
                    "CD측정치", "DC측정치", "(CD-DC)/2", "변위량", "길이환산"]
            for idx, t in enumerate(h_15, 1): 
                ws.cell(row=15, column=idx, value=t).font = Font(bold=True)

            # 단위 행 (Row 16) - 중요!
            units = ["", "(Deg)", "(Deg)", "(Deg)", "(Deg)", "mm", 
                     "(Deg)", "(Deg)", "(Deg)", "(Deg)", "mm"]
            for idx, u in enumerate(units, 1):
                ws.cell(row=16, column=idx, value=u)

            apply_style(ws, 13, 16, 1, 11)
            set_col_width(ws, {'A':15})

            # Data
            for r in range(5):
                rn = 17 + r
                ws.cell(row=rn, column=1, value=f"2025-01-{r+1:02d}")
                for c in range(2, 12):
                    ws.cell(row=rn, column=c, value=round(random.uniform(-0.5, 0.5), 3))
            apply_style(ws, 17, 21, 1, 11)

        wb.save(filepath)
        created_log.append(f"✅ {filename}")

    # ==============================================================================
    # 6. 지표침하계 (SE) - [레벨 야장 형식 완벽 구현]
    # ==============================================================================
    if sensor_counts.get('SE', 0) > 0:
        count = sensor_counts['SE']
        filename = f"지표침하계({project_info}).xlsx"
        filepath = os.path.join(output_dir, filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for i in range(1, count + 1):
            ws = wb.create_sheet(title=f"P.{i}")
            ws['A1'] = "지표침하계 DATA"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:H1')
            
            meta = [
                (4, "■ 계측관리업체 : ㈜지오테크이앤씨"),
                (5, f"■ 현장명 : {project_info}"),
                (6, f"■ 계측기번호 : ST-{i}"),
                (7, "■ 계기종류 : SURFACE SETTLEMENT PIN"),
                (8, "■ 공식 : 측정치 = 기준값 + 후시 - 전시")
            ]
            for r, t in meta:
                ws[f'A{r}'] = t
                ws.merge_cells(f'A{r}:F{r}')
                ws[f'A{r}'].alignment = Alignment(horizontal='left')

            # Headers (12-13행)
            headers = ["날짜", "기준값", "후시(B.S)", "전시(F.S)", "측정치", "변위차", "비고", "CM환산"]
            units = ["", "(M)", "(M)", "(M)", "(M)", "(M)", "", ""]
            
            for idx, t in enumerate(headers, 1):
                ws.cell(row=12, column=idx, value=t).font = Font(bold=True)
                ws.cell(row=13, column=idx, value=units[idx-1])
            
            apply_style(ws, 12, 13, 1, 8)
            set_col_width(ws, {'A':15, 'B':10, 'C':10, 'D':10, 'E':10, 'F':10, 'G':10, 'H':10})

            # Data
            for r in range(5):
                rn = 14 + r
                ws.cell(row=rn, column=1, value=f"2025-01-{r+1:02d}")
                ws.cell(row=rn, column=2, value=10.000)
                ws.cell(row=rn, column=3, value=1.500)
                ws.cell(row=rn, column=4, value=round(1.5 + random.uniform(-0.01, 0.01), 3))
                ws.cell(row=rn, column=5, value=10.0)
                ws.cell(row=rn, column=6, value=0.0)
            apply_style(ws, 14, 18, 1, 8)

        wb.save(filepath)
        created_log.append(f"✅ {filename}")

    return created_log