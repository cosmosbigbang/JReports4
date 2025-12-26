import openpyxl
import os

excel_folder = 'excel'
files = [
    '1. 지중경사계(한남동).xlsx',
    '2. 지하수위계(한남동).xlsx',
    '3. 변형률계(1~7 3단샘플).xlsx',
    '4. 균열측정계(한남동).xlsx',
    '5. 건물경사계 (한남동).xlsx',
    '6. 지표침하계(한남동).xlsx'
]

for filename in files:
    filepath = os.path.join(excel_folder, filename)
    if not os.path.exists(filepath):
        print(f"❌ 파일 없음: {filename}")
        continue
    
    print(f"\n{'='*60}")
    print(f"📁 {filename}")
    print(f"{'='*60}")
    
    try:
        wb = openpyxl.load_workbook(filepath)
        print(f"📋 시트 목록 ({len(wb.sheetnames)}개): {wb.sheetnames[:10]}")  # 처음 10개만
        
        # 첫 시트 분석
        first_sheet = wb[wb.sheetnames[0]]
        print(f"\n🔍 첫 시트 '{first_sheet.title}' 분석:")
        print(f"   사용 영역: {first_sheet.dimensions}")
        
        # 헤더 확인 (첫 3행)
        print(f"\n   📊 첫 3행:")
        for i, row in enumerate(first_sheet.iter_rows(min_row=1, max_row=3, values_only=True), 1):
            non_empty = [cell for cell in row if cell is not None][:15]  # 처음 15컬럼만
            print(f"      Row {i}: {non_empty}")
        
        # 마지막 시트도 확인 (시트가 여러개일 경우)
        if len(wb.sheetnames) > 1:
            last_sheet = wb[wb.sheetnames[-1]]
            print(f"\n   마지막 시트 '{last_sheet.title}':")
            for i, row in enumerate(last_sheet.iter_rows(min_row=1, max_row=2, values_only=True), 1):
                non_empty = [cell for cell in row if cell is not None][:15]
                print(f"      Row {i}: {non_empty}")
        
        wb.close()
        
    except Exception as e:
        print(f"⚠️ 에러: {e}")

print("\n" + "="*60)
print("✅ 분석 완료")
