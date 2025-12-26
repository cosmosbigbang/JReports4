"""
원본 템플릿과 동일한 시트 개수로 생성하여 완벽히 비교
"""
import openpyxl
import os
from excel_template_cloner import create_all_sensor_files, TEMPLATE_MAP

# 1단계: 원본 템플릿의 시트 개수 확인
print("="*60)
print("원본 템플릿 시트 개수 분석")
print("="*60)

template_dir = 'excel'
sensor_counts = {}

for sensor_type, template_file in TEMPLATE_MAP.items():
    template_path = os.path.join(template_dir, template_file)
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    # 시트 개수 계산
    sheet_count = len(wb.sheetnames)
    sensor_counts[sensor_type] = sheet_count
    
    print(f"{sensor_type} ({template_file})")
    print(f"  총 시트 개수: {sheet_count}")
    print(f"  시트 이름: {wb.sheetnames}")
    print()
    
    wb.close()

print("="*60)
print("감지된 계측기 개수:")
print(sensor_counts)
print("="*60)

# 2단계: 동일한 개수로 파일 생성
project_name = "완벽비교_테스트"
print(f"\n생성 시작: {project_name}")
print(f"계측기 개수: {sensor_counts}")

created_files = create_all_sensor_files(sensor_counts, project_name)

print("\n✅ 생성 완료!")
print("="*60)

# 3단계: 파일 크기 비교
print("\n파일 크기 비교:")
print("="*60)

for sensor_type, template_file in TEMPLATE_MAP.items():
    template_path = os.path.join(template_dir, template_file)
    original_size = os.path.getsize(template_path)
    
    if sensor_type in created_files:
        generated_path = created_files[sensor_type]
        generated_size = os.path.getsize(generated_path)
        
        diff_percent = ((generated_size - original_size) / original_size) * 100
        
        print(f"\n{sensor_type} - {template_file}")
        print(f"  원본:     {original_size:,} bytes ({original_size/1024:.1f} KB)")
        print(f"  생성:     {generated_size:,} bytes ({generated_size/1024:.1f} KB)")
        print(f"  차이:     {diff_percent:+.2f}%")

print("\n="*60)

# 4단계: 시트 개수 비교
print("\n시트 개수 비교:")
print("="*60)

for sensor_type, template_file in TEMPLATE_MAP.items():
    template_path = os.path.join(template_dir, template_file)
    
    wb_original = openpyxl.load_workbook(template_path, data_only=False)
    original_sheets = wb_original.sheetnames
    
    if sensor_type in created_files:
        wb_generated = openpyxl.load_workbook(created_files[sensor_type], data_only=False)
        generated_sheets = wb_generated.sheetnames
        
        print(f"\n{sensor_type}:")
        print(f"  원본 시트 수: {len(original_sheets)}")
        print(f"  생성 시트 수: {len(generated_sheets)}")
        print(f"  일치 여부: {'✅ 동일' if len(original_sheets) == len(generated_sheets) else '❌ 다름'}")
        
        if len(original_sheets) != len(generated_sheets):
            print(f"  원본: {original_sheets}")
            print(f"  생성: {generated_sheets}")
        
        wb_generated.close()
    
    wb_original.close()

print("\n" + "="*60)
print("테스트 완료!")
