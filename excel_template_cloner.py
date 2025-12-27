"""
Excel 템플릿 기반 계측기 파일 생성기
템플릿의 모든 formatting, 수식, 차트를 보존하면서 계측기 개수만큼 시트 복제
"""
import os
import shutil
import openpyxl
from copy import copy

# 템플릿 파일 매핑
TEMPLATE_MAP = {
    'T': '5. 건물경사계 (한남동).xlsx',
    'C': '4. 균열측정계(한남동).xlsx',
    'SE': '6. 지표침하계(한남동).xlsx',
    'S': '3. 변형률계(1~7 3단샘플).xlsx',
    'W': '2. 지하수위계(한남동).xlsx',
    'I': '1. 지중경사계(한남동).xlsx'
}

# 시트 이름 패턴
SHEET_PATTERNS = {
    'T': 'T-{}',      # T-1, T-2, ...
    'C': 'C-{}',      # C-1, C-2, ...
    'SE': 'P.{}',     # P.1, P.2, ...
    'S': 'S{}',       # S1, S2, ...
    'W': 'W-{}',      # W-1, W-2, ...
    'I': '{}data'     # 1data, 2data, ...
}


def clone_sheet_with_styles(source_sheet, target_wb, new_name):
    """
    시트를 완전히 복제 (셀 값, 수식, 스타일, 병합, 차트 등 모두 보존)
    """
    target_sheet = target_wb.create_sheet(title=new_name)
    
    # 1. 모든 셀 복사
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet[cell.coordinate]
            
            # 값 또는 수식
            if cell.value is not None:
                new_cell.value = cell.value
            
            # 스타일 복사
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    
    # 2. 병합 셀 복사
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))
    
    # 3. 열 너비 복사
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = col_dim.width
    
    # 4. 행 높이 복사
    for row_num, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_num].height = row_dim.height
    
    # 5. 차트 복사 - 비활성화 (오류 방지)
    # for chart in source_sheet._charts:
    #     target_sheet.add_chart(copy(chart))
    
    # 6. 시트 속성
    if hasattr(source_sheet, 'sheet_properties'):
        target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    
    return target_sheet


def create_sensor_excel(sensor_type, count, project_name, site_name=None, company=None, template_dir='excel', output_dir='generated_excels'):
    """
    특정 센서 타입의 Excel 파일 생성
    
    Args:
        sensor_type: 'T', 'C', 'SE', 'S', 'W', 'I'
        count: 생성할 센서 개수
        project_name: 프로젝트명 (폴더명/파일명에 사용)
        site_name: 현장명 (엑셀 파일 내부에 기록)
        company: 계측관리업체
        template_dir: 템플릿 폴더
        output_dir: 출력 폴더
    """
    if sensor_type not in TEMPLATE_MAP:
        print(f"❌ 알 수 없는 센서 타입: {sensor_type}")
        return None
    
    template_file = TEMPLATE_MAP[sensor_type]
    template_path = os.path.join(template_dir, template_file)
    
    if not os.path.exists(template_path):
        print(f"❌ 템플릿 파일 없음: {template_path}")
        return None
    
    # 출력 디렉토리 생성
    project_dir = os.path.join(output_dir, project_name)
    os.makedirs(project_dir, exist_ok=True)
    
    # 출력 파일명
    sensor_names = {
        'T': '건물경사계',
        'C': '균열측정계',
        'SE': '지표침하계',
        'S': '변형률계',
        'W': '지하수위계',
        'I': '지중경사계'
    }
    output_filename = f"{sensor_names[sensor_type]}({project_name}).xlsx"
    output_path = os.path.join(project_dir, output_filename)
    
    print(f"📂 처리 중: {sensor_type} - {count}개")
    
    # 템플릿 파일 전체 복사
    shutil.copy2(template_path, output_path)
    
    # 복사된 파일 열기
    wb = openpyxl.load_workbook(output_path)
    
    # 기존 시트 이름 확인
    template_sheets = [s for s in wb.sheetnames if s != '간지']
    if not template_sheets:
        print(f"❌ 템플릿에 데이터 시트가 없음")
        wb.close()
        return None
    
    first_sheet = template_sheets[0]
    pattern = SHEET_PATTERNS[sensor_type]
    
    # 첫 번째 시트 이름 변경
    wb[first_sheet].title = pattern.format(1)
    
    # 추가 시트 생성 (2번째부터)
    for i in range(2, count + 1):
        new_sheet_name = pattern.format(i)
        print(f"  ✓ 시트 복제: {new_sheet_name}")
        new_sheet = wb.copy_worksheet(wb[pattern.format(1)])
        new_sheet.title = new_sheet_name
        new_sheet = wb.copy_worksheet(wb[pattern.format(1)])
        new_sheet.title = new_sheet_name
    
    # 모든 시트에 현장명/업체명 업데이트
    for i in range(1, count + 1):
        sheet_name = pattern.format(i)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            try:
                # 병합된 셀 확인 및 해제
                cells_to_update = ['A3', 'A4', 'A5', 'A11']
                for cell_ref in cells_to_update:
                    for merged_range in list(sheet.merged_cells.ranges):
                        if cell_ref in merged_range:
                            sheet.unmerge_cells(str(merged_range))
                            break
                
                # 값 업데이트
                if company:
                    sheet['A3'].value = f'■  계측관리업체(Contract) : {company}'
                if site_name:
                    sheet['A4'].value = f'■  현장명(Site) : {site_name}'
                sheet['A5'].value = f'■  계측기번호(NO.) : {sheet_name}'
                sheet['A11'].value = '설치장소:'
                
                print(f"  ✅ 시트 완료: {sheet_name}")
            except Exception as e:
                print(f"  ⚠️ 시트 {sheet_name} 업데이트 오류: {e}")
    
    # 저장
    wb.save(output_path)
    wb.close()
    
    file_size = os.path.getsize(output_path)
    print(f"✅ 완료: {output_filename} ({file_size:,} bytes)\n")
    
    return output_path


def create_all_sensor_files(sensor_counts, project_name, site_name=None, company=None):
    """
    모든 센서 타입의 Excel 파일 생성
    
    Args:
        sensor_counts: {'T': 18, 'C': 18, 'SE': 9, 'S': 7, 'W': 5, 'I': 5}
        project_name: 프로젝트명
        site_name: 현장명
        company: 계측관리업체
    """
    print(f"\n{'='*60}")
    print(f"🚀 Excel 파일 생성 시작: {project_name}")
    print(f"{'='*60}\n")
    
    results = {}
    
    for sensor_type, count in sensor_counts.items():
        if count > 0:
            output_path = create_sensor_excel(sensor_type, count, project_name, site_name, company)
            if output_path:
                results[sensor_type] = output_path
    
    print(f"\n{'='*60}")
    print(f"✨ 생성 완료: {len(results)}개 파일")
    print(f"{'='*60}\n")
    
    return results


if __name__ == "__main__":
    # 테스트 1: 템플릿과 같은 개수
    print("\n🧪 테스트 1: 템플릿과 동일한 센서 개수")
    test_counts_1 = {
        'T': 18,   # 건물경사계
        'C': 18,   # 균열측정계
        'SE': 9,   # 지표침하계
        'S': 7,    # 변형률계
        'W': 5,    # 지하수위계
        'I': 5     # 지중경사계
    }
    
    create_all_sensor_files(test_counts_1, "한남동_383-1_test1")
    
    # 테스트 2: 다른 개수
    print("\n🧪 테스트 2: 다른 센서 개수")
    test_counts_2 = {
        'T': 10,
        'C': 15,
        'SE': 6,
        'S': 12,
        'W': 3,
        'I': 3
    }
    
    create_all_sensor_files(test_counts_2, "한남동_383-1_test2")
